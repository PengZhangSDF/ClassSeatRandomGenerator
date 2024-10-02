from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import os
from PyQt5.QtWidgets import QMessageBox, QFileDialog

def export_seating_arrangement(seats, boys, girls, parent=None):
    # 获取当前日期
    today = datetime.now()
    month = today.month
    day = today.day

    # 生成文件名，递增数字
    base_filename = f"学生座位{month}月{day}日"
    file_index = 1
    filename = f"{base_filename}({file_index}).xlsx"

    # 检查文件是否已存在，递增数字
    while os.path.exists(filename):
        file_index += 1
        filename = f"{base_filename}({file_index}).xlsx"

    # 打开文件对话框让用户选择保存位置
    save_path, _ = QFileDialog.getSaveFileName(parent, "保存座位表", filename, "Excel Files (*.xlsx)")
    if not save_path:  # 如果用户取消了对话框
        return

    # 创建一个新的工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Seating Arrangement"

    # 创建一个字典来映射姓名到颜色
    name_color_map = {}
    for name, _, label in boys:
        name_color_map[name] = "lightblue"  # 男生对应的颜色
    for name, _, label in girls:
        name_color_map[name] = "lightyellow"  # 女生对应的颜色

    # 遍历座位并填充 Excel
    for row in range(len(seats)):
        for col in range(len(seats[row])):
            seat = seats[row][col]
            seat_text = seat.text()

            # 只记录有姓名的座位
            if seat_text:  
                excel_column = chr(col + 65)  # 将列索引转换为 Excel 字母（0 -> A, 1 -> B, ...）
                excel_row = row + 1  # Excel 行从 1 开始
                cell_position = f"{excel_column}{excel_row}"  # 生成 Excel 单元格位置

                # 设置单元格的值和背景颜色
                worksheet[cell_position] = seat_text  # 设置单元格的姓名
                color = name_color_map.get(seat_text, "white")  # 获取颜色
                if color == "lightblue":
                    worksheet[cell_position].fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                elif color == "lightyellow":
                    worksheet[cell_position].fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

                # 调试信息
                print(f"Exporting: {seat_text} to {cell_position} with color {color}")

    # 保存工作簿
    workbook.save(save_path)
    QMessageBox.information(parent, "保存成功", f"座位表已保存到 {save_path}。")